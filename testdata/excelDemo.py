import openpyxl
Dict={}
workbook = openpyxl.load_workbook("/Users/ssaguturu/Downloads/Python_excel_Demo.xlsx")
#workbook = openpyxl.load_workbook("https://docs.google.com/spreadsheets/d/1TUTYAhgqVMXl7IXuSMgDUlrxl8eK2a2hwJRUKE8N5Yg/edit#gid=0")
sheet = workbook.active
#cell = sheet.cell(row=2,column=2)
#print(cell.value)
#sheet.cell(row=6,column=2).value = "abc"
#sheet.cell(row=7,column=2).value = "abcd"
#print(sheet.cell(row=6,column=2).value)
#print(sheet.cell(row=7,column=2).value)
#print(sheet.max_row)
#print(sheet.max_column)
#print(sheet['B5'].value)
for i in range(1,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):

        #if sheet.cell(row=i,column=1).value == "TestCase1":

            print(sheet.cell(row=i, column=j).value)

            Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
print(Dict)


workbook.save('workbook')
workbook.close()

