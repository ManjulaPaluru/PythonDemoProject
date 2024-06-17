import openpyxl

from testdata.excelDemo import sheet


class HomePageData:
# here testdata sets are wrapped into square brackets ,it should be wrapped under list

    test_HomePage_data = [{"firstname" :"manju","lastname":"paluru","gender":"Male"},{"firstname" :"anju","lastname":"aluru","gender":"Female"}]

    @staticmethod
    def get_text_data(test_case_name):
        Dict = {}
        workbook = openpyxl.load_workbook("/Users/ssaguturu/Downloads/Python_excel_Demo (1).xlsx")
        sheet = workbook.active
        for i in range(1,sheet.max_row+1): # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2,sheet.max_column+1):  # to get columns

                    Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
        return[Dict]
# wrapping Dictionary data into list.




